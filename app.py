import base64
import os
import sys
import threading
import time

# --- 1. PERFORMANCE & PACKAGING CONFIG ---
os.environ["SETTINGS_SYNC"] = "False"
os.environ["YOLO_VERBOSE"] = "False"
os.environ["OPENCV_VIDEOIO_MSMF_ENABLE_HW_TRANSFORMS"] = "0"
os.environ.setdefault("OMP_NUM_THREADS", str(max(1, min(4, (os.cpu_count() or 2) - 1))))
os.environ.setdefault("MKL_NUM_THREADS", os.environ["OMP_NUM_THREADS"])

import cv2
import flet as ft
import multiprocessing
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook

try:
    cv2.setNumThreads(1)
    cv2.ocl.setUseOpenCL(False)
except Exception:
    pass


def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller."""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def get_desktop_path():
    return os.path.join(os.path.expanduser("~"), "Desktop")


EXCEL_FILE = os.path.join(get_desktop_path(), "acne_history.xlsx")
IMAGE_SAVE_DIR = os.path.join(get_desktop_path(), "Acne_Annotated_Images")
MODEL_PATH = resource_path("best.pt")

PREVIEW_SIZE = (850, 500)
CAMERA_SIZE = (960, 540)
MODEL_IMAGE_SIZE = 640
CONFIDENCE_THRESHOLD = 0.03
FALLBACK_CONFIDENCE_THRESHOLD = 0.01
STREAM_INTERVAL_SEC = 1 / 24


class AcneKioskFletApp:
    def __init__(self, page):
        self.page = page
        self.model = None
        self.preview_cap = None
        self.preview_running = False
        self.inference_running = False
        self.current_frame = None
        self.last_annotated_frame = None
        self.frame_lock = threading.Lock()

        self.patient_entry = None
        self.status_label = None
        self.acne_count_label = None
        self.preview_image = None
        self.preview_placeholder = None
        self.history_column = None

        self.build_ui()
        self.page.run_thread(self._lazy_load_engine)
        self.load_history()

    def build_ui(self):
        self.page.title = "AI Acne Detection (PyTorch CPU)"
        self.page.bgcolor = "#020617"
        self.page.padding = 0
        self.page.window_width = 1200
        self.page.window_height = 820
        self.page.window_min_width = 980
        self.page.window_min_height = 720

        self.patient_entry = ft.TextField(
            hint_text="Patient Name...",
            width=350,
            height=45,
            border_color="#334155",
            focused_border_color="#3b82f6",
            bgcolor="#0f172a",
            color="#e2e8f0",
        )
        self.status_label = ft.Text("Initializing...", color="#f59e0b", weight=ft.FontWeight.BOLD, size=13)
        self.acne_count_label = ft.Text("0", color="#3b82f6", weight=ft.FontWeight.BOLD, size=32)

        count_card = ft.Container(
            width=150,
            bgcolor="#1e293b",
            border_radius=8,
            padding=ft.Padding(12, 8, 12, 8),
            content=ft.Column(
                controls=[
                    ft.Text("ACNE COUNT", size=10, color="#cbd5e1"),
                    self.acne_count_label,
                ],
                spacing=0,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                alignment=ft.MainAxisAlignment.CENTER,
            ),
        )

        header = ft.Container(
            height=100,
            bgcolor="#111827",
            padding=ft.Padding(20, 16, 20, 16),
            content=ft.Row(
                controls=[
                    self.patient_entry,
                    self.status_label,
                    ft.Container(expand=True),
                    count_card,
                ],
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
            ),
        )

        self.preview_image = ft.Image(
            src="",
            width=PREVIEW_SIZE[0],
            height=PREVIEW_SIZE[1],
            fit=ft.BoxFit.CONTAIN,
            gapless_playback=True,
            visible=False,
        )
        self.preview_placeholder = ft.Text("Camera Offline", size=20, color="#cbd5e1")
        display_frame = ft.Container(
            expand=True,
            bgcolor="#0f172a",
            border_radius=8,
            alignment=ft.Alignment(0, 0),
            content=ft.Stack(
                controls=[self.preview_image, self.preview_placeholder],
                alignment=ft.Alignment(0, 0),
            ),
        )

        self.history_column = ft.Column(spacing=8, scroll=ft.ScrollMode.AUTO, expand=True)
        sidebar = ft.Container(
            width=300,
            bgcolor="#111827",
            border_radius=8,
            padding=10,
            content=ft.Column(
                controls=[
                    ft.Text("Recent History", size=16, weight=ft.FontWeight.BOLD, color="#e2e8f0"),
                    ft.Container(
                        expand=True,
                        bgcolor="#020617",
                        border_radius=8,
                        padding=10,
                        content=self.history_column,
                    ),
                ],
                expand=True,
            ),
        )

        content = ft.Row(
            controls=[
                display_frame,
                sidebar,
            ],
            expand=True,
            spacing=10,
        )

        buttons = ft.Container(
            height=80,
            padding=ft.Padding(10, 14, 10, 14),
            bgcolor="#020617",
            content=ft.Row(
                controls=[
                    ft.Button("Open Camera", height=45, on_click=self.open_camera),
                    ft.Button(
                        "Capture & Analyze",
                        height=45,
                        bgcolor="#2563eb",
                        color="#ffffff",
                        on_click=self.capture_image,
                    ),
                    ft.Button("Save Record", height=45, on_click=self.save_annotated_image),
                    ft.Container(expand=True),
                    ft.Button(
                        "Exit",
                        height=45,
                        bgcolor="#475569",
                        color="#ffffff",
                        on_click=self.exit_app,
                    ),
                ],
                spacing=10,
            ),
        )

        self.page.add(
            ft.Column(
                controls=[
                    header,
                    ft.Container(content=content, expand=True, padding=10),
                    buttons,
                ],
                expand=True,
                spacing=0,
            )
        )
        self.page.update()

    def _lazy_load_engine(self):
        try:
            import torch
            from ultralytics import YOLO

            torch.set_num_threads(max(1, min(4, (os.cpu_count() or 2) - 1)))
            torch.set_num_interop_threads(1)

            if not os.path.exists(MODEL_PATH):
                raise FileNotFoundError(f"Model not found at: {MODEL_PATH}")

            loaded_model = YOLO(MODEL_PATH)
            loaded_model.to("cpu")
            try:
                loaded_model.fuse()
            except Exception:
                pass
            loaded_model.model.eval()

            with torch.inference_mode():
                loaded_model.predict(
                    np.zeros((MODEL_IMAGE_SIZE, MODEL_IMAGE_SIZE, 3), dtype=np.uint8),
                    imgsz=MODEL_IMAGE_SIZE,
                    device="cpu",
                    verbose=False,
                )

            self.model = loaded_model
            self._set_status("System Ready", "#22c55e")
        except Exception as err:
            self._set_status(f"Error: {err}", "red")

    def open_camera(self, _=None):
        if self.preview_running:
            return

        self.preview_cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
        self.preview_cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*"MJPG"))
        self.preview_cap.set(cv2.CAP_PROP_FRAME_WIDTH, CAMERA_SIZE[0])
        self.preview_cap.set(cv2.CAP_PROP_FRAME_HEIGHT, CAMERA_SIZE[1])
        self.preview_cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)

        if not self.preview_cap.isOpened():
            self._set_status("Camera Error", "red")
            return

        self.preview_running = True
        self._set_status("Camera Active", "#3b82f6")
        self.page.run_thread(self._stream_worker)

    def _stream_worker(self):
        while self.preview_running:
            ret, frame = self.preview_cap.read()
            if not ret:
                break

            with self.frame_lock:
                self.current_frame = frame

            self._show_frame(frame)
            time.sleep(STREAM_INTERVAL_SEC)

        if self.preview_running:
            self.preview_running = False
            self._set_status("Camera Feed Stopped", "red")

    def capture_image(self, _=None):
        if self.inference_running or not self.preview_running or self.model is None:
            return

        with self.frame_lock:
            if self.current_frame is None:
                return
            snap = self.current_frame.copy()

        self.preview_running = False
        if self.preview_cap:
            self.preview_cap.release()

        self.inference_running = True
        self._set_status("Analyzing...", "#f59e0b")
        self.page.run_thread(self._analyze_frame_worker, snap)

    def _analyze_frame_worker(self, frame):
        try:
            import torch

            with torch.inference_mode():
                results = self.model.predict(
                    frame,
                    conf=CONFIDENCE_THRESHOLD,
                    imgsz=MODEL_IMAGE_SIZE,
                    device="cpu",
                    verbose=False,
                )[0]
                if len(results.boxes) == 0:
                    results = self.model.predict(
                        frame,
                        conf=FALLBACK_CONFIDENCE_THRESHOLD,
                        imgsz=MODEL_IMAGE_SIZE,
                        device="cpu",
                        verbose=False,
                    )[0]

            annotated = results.plot()
            count = len(results.boxes)
            self._show_analysis_result(annotated, count)
        except Exception as err:
            self._set_status(f"Error: {err}", "red")
            self.inference_running = False

    def _show_analysis_result(self, annotated_frame, count):
        self.last_annotated_frame = annotated_frame
        self.acne_count_label.value = str(count)
        self._show_frame(annotated_frame)
        self.save_to_excel(count)
        self._set_status("Analysis Complete", "#22c55e")
        self.inference_running = False

    def _show_frame(self, frame):
        image_base64, preview_size = self._frame_to_base64(frame)
        self.preview_image.src = image_base64
        self.preview_image.width = preview_size[0]
        self.preview_image.height = preview_size[1]
        self.preview_image.visible = True
        self.preview_placeholder.visible = False
        self.page.update()

    def _frame_to_base64(self, frame):
        height, width = frame.shape[:2]
        max_width, max_height = PREVIEW_SIZE
        scale = min(max_width / width, max_height / height, 1.0)
        preview_size = (int(width * scale), int(height * scale))
        resized = cv2.resize(frame, preview_size, interpolation=cv2.INTER_AREA)
        ok, buffer = cv2.imencode(".jpg", resized, [int(cv2.IMWRITE_JPEG_QUALITY), 82])
        if not ok:
            raise RuntimeError("Could not encode camera frame")
        return base64.b64encode(buffer).decode("ascii"), preview_size

    def save_annotated_image(self, _=None):
        if self.last_annotated_frame is None:
            return

        os.makedirs(IMAGE_SAVE_DIR, exist_ok=True)
        name = self.patient_entry.value or "Unknown"
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        cv2.imwrite(os.path.join(IMAGE_SAVE_DIR, f"{name}_{ts}.jpg"), self.last_annotated_frame)
        self._set_status(f"Saved {ts}", "#22c55e")

    def save_to_excel(self, count):
        row = [datetime.now().strftime("%Y-%m-%d %H:%M"), self.patient_entry.value or "Unknown", count]
        try:
            if os.path.exists(EXCEL_FILE):
                workbook = load_workbook(EXCEL_FILE)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["Date", "Patient", "Count"])

            sheet.append(row)
            workbook.save(EXCEL_FILE)
        except Exception as err:
            self._set_status(f"History Error: {err}", "red")
        self.load_history()

    def load_history(self):
        self.history_column.controls.clear()
        if not os.path.exists(EXCEL_FILE):
            self.page.update()
            return

        try:
            workbook = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(min_row=2, values_only=True))[-15:]
            workbook.close()

            for date, patient, count in reversed(rows):
                self.history_column.controls.append(
                    ft.Container(
                        border=ft.Border(bottom=ft.BorderSide(1, "#1e293b")),
                        padding=ft.Padding(0, 0, 0, 8),
                        content=ft.Column(
                            controls=[
                                ft.Text(str(date), size=12, color="#94a3b8"),
                                ft.Text(f"{patient}: {count} spots", size=12, color="#e2e8f0"),
                            ],
                            spacing=2,
                        ),
                    )
                )
        except Exception as err:
            self._set_status(f"History Error: {err}", "red")
        self.page.update()

    def _set_status(self, text, color):
        self.status_label.value = text
        self.status_label.color = color
        self.page.update()

    def exit_app(self, _=None):
        self.preview_running = False
        if self.preview_cap:
            self.preview_cap.release()
        if hasattr(self.page, "window") and hasattr(self.page.window, "close"):
            self.page.window.close()
        else:
            self.page.window_close()


def main(page):
    AcneKioskFletApp(page)


if __name__ == "__main__":
    multiprocessing.freeze_support()
    ft.run(main)
