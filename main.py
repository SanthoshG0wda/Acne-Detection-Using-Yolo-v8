import os
import sys
import threading
import time

# --- 1. PERFORMANCE & PACKAGING CONFIG ---
os.environ["SETTINGS_SYNC"] = "False"
os.environ["YOLO_VERBOSE"] = "False"
os.environ["OPENCV_VIDEOIO_MSMF_ENABLE_HW_TRANSFORMS"] = "0"
os.environ.setdefault("OMP_NUM_THREADS", str(max(1, min(4, (os.cpu_count() or 2) - 1))))
os.environ.setdefault("MKL_NUM_THREADS", os.environ["OMP_NUM_THREADS"])

import cv2
import customtkinter as ctk
import multiprocessing
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook
from PIL import Image

try:
    cv2.setNumThreads(1)
    cv2.ocl.setUseOpenCL(False)
except Exception:
    pass


def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller."""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def get_desktop_path():
    return os.path.join(os.path.expanduser("~"), "Desktop")


EXCEL_FILE = os.path.join(get_desktop_path(), "acne_history.xlsx")
IMAGE_SAVE_DIR = os.path.join(get_desktop_path(), "Acne_Annotated_Images")
MODEL_PATH = resource_path("best.pt")

PREVIEW_SIZE = (850, 500)
CAMERA_SIZE = (960, 540)
MODEL_IMAGE_SIZE = 640
CONFIDENCE_THRESHOLD = 0.03
FALLBACK_CONFIDENCE_THRESHOLD = 0.01
PREVIEW_INTERVAL_MS = 50
STREAM_INTERVAL_SEC = 1 / 24


class AcneKioskApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("AI Acne Detection (PyTorch CPU)")
        self.geometry("1200x820")

        self.model = None
        self.preview_cap = None
        self.preview_running = False
        self.inference_running = False
        self.current_frame = None
        self.last_annotated_frame = None
        self.frame_lock = threading.Lock()

        self.build_ui()
        threading.Thread(target=self._lazy_load_engine, daemon=True).start()
        self.load_history()

    def _lazy_load_engine(self):
        try:
            import torch
            from ultralytics import YOLO

            torch.set_num_threads(max(1, min(4, (os.cpu_count() or 2) - 1)))
            torch.set_num_interop_threads(1)

            if not os.path.exists(MODEL_PATH):
                raise FileNotFoundError(f"Model not found at: {MODEL_PATH}")

            loaded_model = YOLO(MODEL_PATH)
            loaded_model.to("cpu")
            try:
                loaded_model.fuse()
            except Exception:
                pass
            loaded_model.model.eval()

            with torch.inference_mode():
                loaded_model.predict(
                    np.zeros((MODEL_IMAGE_SIZE, MODEL_IMAGE_SIZE, 3), dtype=np.uint8),
                    imgsz=MODEL_IMAGE_SIZE,
                    device="cpu",
                    verbose=False,
                )

            self.model = loaded_model
            self.after(0, self._set_status_ready)
        except Exception as err:
            error_msg = str(err)
            self.after(0, lambda: self.status_label.configure(text=f"Error: {error_msg}", text_color="red"))

    def _set_status_ready(self):
        self.status_label.configure(text="System Ready", text_color="#22c55e")

    def build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=0)
        self.grid_rowconfigure(1, weight=1)

        self.top_frame = ctk.CTkFrame(self, height=100, corner_radius=0)
        self.top_frame.grid(row=0, column=0, columnspan=2, sticky="ew")

        self.patient_entry = ctk.CTkEntry(self.top_frame, placeholder_text="Patient Name...", width=350, height=45)
        self.patient_entry.pack(side="left", padx=20, pady=20)

        self.status_label = ctk.CTkLabel(
            self.top_frame,
            text="Initializing...",
            text_color="#f59e0b",
            font=("Arial", 13, "bold"),
        )
        self.status_label.pack(side="left", padx=10)

        self.count_card = ctk.CTkFrame(self.top_frame, fg_color="#1e293b", width=150)
        self.count_card.pack(side="right", padx=20, pady=10)
        ctk.CTkLabel(self.count_card, text="ACNE COUNT", font=("Arial", 10)).pack(padx=10)
        self.acne_count_label = ctk.CTkLabel(
            self.count_card,
            text="0",
            font=("Arial", 32, "bold"),
            text_color="#3b82f6",
        )
        self.acne_count_label.pack(padx=10, pady=(0, 5))

        self.display_frame = ctk.CTkFrame(self, fg_color="#0f172a")
        self.display_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        self.image_label = ctk.CTkLabel(self.display_frame, text="Camera Offline", font=("Arial", 20))
        self.image_label.pack(expand=True, fill="both")

        self.sidebar = ctk.CTkFrame(self, width=300)
        self.sidebar.grid(row=1, column=1, rowspan=2, sticky="ns", padx=10, pady=10)
        ctk.CTkLabel(self.sidebar, text="Recent History", font=("Arial", 16, "bold")).pack(pady=10)
        self.history_box = ctk.CTkTextbox(self.sidebar, width=280, font=("Consolas", 12))
        self.history_box.pack(expand=True, fill="both", padx=10, pady=10)

        self.btn_frame = ctk.CTkFrame(self, height=80)
        self.btn_frame.grid(row=2, column=0, sticky="ew", padx=10, pady=10)

        ctk.CTkButton(self.btn_frame, text="Open Camera", command=self.open_camera, height=45).pack(side="left", padx=10)
        ctk.CTkButton(
            self.btn_frame,
            text="Capture & Analyze",
            command=self.capture_image,
            fg_color="#2563eb",
            height=45,
        ).pack(side="left", padx=10)
        ctk.CTkButton(self.btn_frame, text="Save Record", command=self.save_annotated_image, height=45).pack(
            side="left",
            padx=10,
        )
        ctk.CTkButton(self.btn_frame, text="Exit", fg_color="#475569", command=self.exit_app, height=45).pack(
            side="right",
            padx=10,
        )

    def open_camera(self):
        if self.preview_running:
            return

        self.preview_cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
        self.preview_cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*"MJPG"))
        self.preview_cap.set(cv2.CAP_PROP_FRAME_WIDTH, CAMERA_SIZE[0])
        self.preview_cap.set(cv2.CAP_PROP_FRAME_HEIGHT, CAMERA_SIZE[1])
        self.preview_cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)

        if not self.preview_cap.isOpened():
            self.status_label.configure(text="Camera Error", text_color="red")
            return

        self.preview_running = True
        self.status_label.configure(text="Camera Active", text_color="#3b82f6")
        threading.Thread(target=self._stream_worker, daemon=True).start()
        self._update_ui_frame()

    def _stream_worker(self):
        while self.preview_running:
            ret, frame = self.preview_cap.read()
            if not ret:
                break
            with self.frame_lock:
                self.current_frame = frame
            time.sleep(STREAM_INTERVAL_SEC)

    def _update_ui_frame(self):
        if not self.preview_running:
            return

        with self.frame_lock:
            frame = self.current_frame.copy() if self.current_frame is not None else None

        if frame is not None:
            preview, preview_size = self._frame_to_preview(frame)
            img = ctk.CTkImage(Image.fromarray(preview), size=preview_size)
            self.image_label.configure(image=img, text="")

        self.after(PREVIEW_INTERVAL_MS, self._update_ui_frame)

    def capture_image(self):
        if self.inference_running or not self.preview_running or self.model is None:
            return

        with self.frame_lock:
            if self.current_frame is None:
                return
            snap = self.current_frame.copy()

        self.preview_running = False
        if self.preview_cap:
            self.preview_cap.release()

        self.inference_running = True
        self.status_label.configure(text="Analyzing...", text_color="#f59e0b")
        threading.Thread(target=self._analyze_frame_worker, args=(snap,), daemon=True).start()

    def _analyze_frame_worker(self, frame):
        try:
            import torch

            with torch.inference_mode():
                results = self.model.predict(
                    frame,
                    conf=CONFIDENCE_THRESHOLD,
                    imgsz=MODEL_IMAGE_SIZE,
                    device="cpu",
                    verbose=False,
                )[0]
                if len(results.boxes) == 0:
                    results = self.model.predict(
                        frame,
                        conf=FALLBACK_CONFIDENCE_THRESHOLD,
                        imgsz=MODEL_IMAGE_SIZE,
                        device="cpu",
                        verbose=False,
                    )[0]

            annotated = results.plot()
            count = len(results.boxes)
            self.after(0, lambda: self._show_analysis_result(annotated, count))
        except Exception as err:
            error_msg = str(err)
            self.after(0, lambda: self.status_label.configure(text=f"Error: {error_msg}", text_color="red"))
            self.after(0, self._finish_inference)

    def _show_analysis_result(self, annotated_frame, count):
        self.last_annotated_frame = annotated_frame
        self.acne_count_label.configure(text=str(count))

        res_rgb, preview_size = self._frame_to_preview(self.last_annotated_frame)
        img = ctk.CTkImage(Image.fromarray(res_rgb), size=preview_size)
        self.image_label.configure(image=img)

        self.save_to_excel(count)
        self.status_label.configure(text="Analysis Complete", text_color="#22c55e")
        self._finish_inference()

    def _finish_inference(self):
        self.inference_running = False

    def _frame_to_preview(self, frame):
        height, width = frame.shape[:2]
        max_width, max_height = PREVIEW_SIZE
        scale = min(max_width / width, max_height / height, 1.0)
        preview_size = (int(width * scale), int(height * scale))
        resized = cv2.resize(frame, preview_size, interpolation=cv2.INTER_AREA)
        return cv2.cvtColor(resized, cv2.COLOR_BGR2RGB), preview_size

    def save_annotated_image(self):
        if self.last_annotated_frame is None:
            return

        os.makedirs(IMAGE_SAVE_DIR, exist_ok=True)
        name = self.patient_entry.get() or "Unknown"
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        cv2.imwrite(os.path.join(IMAGE_SAVE_DIR, f"{name}_{ts}.jpg"), self.last_annotated_frame)
        self.status_label.configure(text=f"Saved {ts}", text_color="#22c55e")

    def save_to_excel(self, count):
        row = [datetime.now().strftime("%Y-%m-%d %H:%M"), self.patient_entry.get() or "Unknown", count]
        try:
            if os.path.exists(EXCEL_FILE):
                workbook = load_workbook(EXCEL_FILE)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["Date", "Patient", "Count"])

            sheet.append(row)
            workbook.save(EXCEL_FILE)
        except Exception as err:
            self.status_label.configure(text=f"History Error: {err}", text_color="red")
        self.load_history()

    def load_history(self):
        self.history_box.delete("1.0", "end")
        if not os.path.exists(EXCEL_FILE):
            return

        try:
            workbook = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(min_row=2, values_only=True))[-15:]
            workbook.close()
            for date, patient, count in reversed(rows):
                self.history_box.insert("end", f"{date}\n{patient}: {count} spots\n{'-' * 20}\n")
        except Exception as err:
            self.status_label.configure(text=f"History Error: {err}", text_color="red")

    def exit_app(self):
        self.preview_running = False
        if self.preview_cap:
            self.preview_cap.release()
        self.destroy()


if __name__ == "__main__":
    multiprocessing.freeze_support()
    app = AcneKioskApp()
    app.mainloop()
